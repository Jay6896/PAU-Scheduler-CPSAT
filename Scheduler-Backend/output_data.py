import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from collections import defaultdict
from datetime import datetime
import io

class TimetableExporter:
    def __init__(self):
        self.saved_data_path = os.path.join(os.path.dirname(__file__), 'data', 'timetable_data.json')
        self.output_dir = os.path.join(os.path.dirname(__file__), 'output_data')
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Load course, room, and faculty data
        self.course_data = self.load_course_data()
        self.room_data = self.load_room_data()
        self.faculty_data = self.load_faculty_data()
        
        # Keywords to identify SST (engineering) groups
        self.sst_keywords = [
            'engineering', 'eng', 'computer science', 'software engineering', 'data science',
            'mechatronics', 'electrical', 'mechanical', 'csc', 'sen', 'data', 'ds'
        ]
        
        # Time slots
        self.time_slots = [
            "08:30-09:30", "09:30-10:30", "10:30-11:30", "11:30-12:30", "12:30-13:30",
            "13:30-14:30", "14:30-15:30", "15:30-16:30", "16:30-17:30", "17:30-18:30"
        ]
        
        # Days of the week
        self.days = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

    def load_course_data(self):
        """Load course data from JSON file"""
        try:
            course_path = os.path.join(os.path.dirname(__file__), 'data', 'course-data.json')
            with open(course_path, 'r', encoding='utf-8') as f:
                courses = json.load(f)
            
            # Create a mapping from course code to course info
            course_map = {}
            for course in courses:
                course_map[course['code']] = course
            
            return course_map
        except Exception as e:
            print(f"Error loading course data: {e}")
            return {}

    def load_room_data(self):
        """Load room data from JSON file"""
        try:
            room_path = os.path.join(os.path.dirname(__file__), 'data', 'rooms-data.json')
            with open(room_path, 'r', encoding='utf-8') as f:
                rooms = json.load(f)
            
            # Create a mapping from room name to room info
            room_map = {}
            for room in rooms:
                room_map[room['name']] = room
            
            return room_map
        except Exception as e:
            print(f"Error loading room data: {e}")
            return {}

    def load_faculty_data(self):
        """Load faculty data from JSON file"""
        try:
            faculty_path = os.path.join(os.path.dirname(__file__), 'data', 'faculty-data.json')
            with open(faculty_path, 'r', encoding='utf-8') as f:
                faculty = json.load(f)
            
            # Create a mapping from faculty ID to faculty info
            faculty_map = {}
            for fac in faculty:
                faculty_map[fac['id']] = fac  # Use 'id' field instead of 'faculty_id'
            
            return faculty_map
        except Exception as e:
            print(f"Error loading faculty data: {e}")
            return {}

    def load_saved_timetable_data(self):
        """Load the latest saved timetable data"""
        try:
            if os.path.exists(self.saved_data_path):
                with open(self.saved_data_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # FIX: Check for the new dictionary format containing 'timetables'
                if isinstance(data, dict) and 'timetables' in data:
                    print(f"Loaded saved timetable data (new format): {len(data['timetables'])} groups")
                    return data['timetables'] # Return only the list of timetables
                else:
                    # Handle old format for backward compatibility
                    print(f"Loaded saved timetable data (old format): {len(data)} groups")
                    return data
            else:
                print("No saved timetable data found")
                return None
        except Exception as e:
            print(f"Error loading saved data: {e}")
            return None

    def load_fresh_optimization_data(self):
        """Load fresh DE optimization data if available"""
        fresh_data_path = os.path.join(os.path.dirname(__file__), 'data', 'fresh_timetable_data.json')
        try:
            if os.path.exists(fresh_data_path):
                with open(fresh_data_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                print(f"Loaded fresh optimization data: {len(data)} groups")
                return data
            else:
                print("No fresh optimization data found")
                return None
        except Exception as e:
            print(f"Error loading fresh optimization data: {e}")
            return None

    def get_timetable_data(self):
        """Get timetable data - prefer saved data, fallback to fresh optimization data"""
        # First try to load saved data (user modifications)
        data = self.load_saved_timetable_data()
        if data:
            return data
        
        # If no saved data, try to load fresh optimization data
        data = self.load_fresh_optimization_data()
        if data:
            return data
        
        return None

    def _normalize_building(self, raw_building) -> str:
        if raw_building is None:
            return ""
        b = str(raw_building).strip().upper()
        if not b:
            return ""
        if b == 'SST' or b.startswith('SST'):
            return 'SST'
        if b == 'TYD' or b.startswith('TYD'):
            return 'TYD'
        return ""

    def is_sst_group(self, group):
        """Building-first SST check; keyword matching only as fallback when building is invalid."""
        # group may be a string (older format) or a dict like {'name': ..., 'building': ...}
        if isinstance(group, dict):
            nb = self._normalize_building(group.get('building') or group.get('effective_building'))
            if nb in {'SST', 'TYD'}:
                return nb == 'SST'
            name = str(group.get('name') or '')
        else:
            name = str(group or '')

        name_lower = name.lower()
        return any(keyword in name_lower for keyword in self.sst_keywords)

    def extract_main_program_name(self, group_name):
        """Extract the main program name from student group name"""
        # Remove year information and stream information
        # Examples: "Computer Science - Year 1" -> "Computer Science"
        #          "Electrical Engineering - Year 2/Stream 1" -> "Electrical Engineering"
        
        # First check for hyphen
        if ' - ' in group_name:
            main_name = group_name.split(' - ')[0]
        else:
            # If no hyphen, look for the word "year" (case insensitive)
            match = re.search(r'\s+year\s+\d+', group_name, re.IGNORECASE)
            if match:
                main_name = group_name[:match.start()]
            else:
                main_name = group_name
        
        # Remove any stream information that might be in the main name
        main_name = re.sub(r'/Stream \d+', '', main_name)
        main_name = re.sub(r'Stream \d+', '', main_name)
        
        return main_name.strip()

    def split_lecturer_names(self, faculty_raw):
        """Split multi-lecturer strings only on ',' or '/', never on spaces."""
        s = str(faculty_raw or '').strip()
        if not s:
            return []

        def _clean_one(name: str) -> str:
            # Strip email/extra info in parentheses e.g. "Dr X (x@pau.edu)"
            return re.sub(r'\s*\([^)]*\)\s*', '', str(name or '')).strip()

        if ',' not in s and '/' not in s:
            one = _clean_one(s)
            return [one] if one else []

        parts = [p.strip() for p in re.split(r'\s*[,/]\s*', s) if p and p.strip()]
        cleaned = []
        for p in parts:
            c = _clean_one(p)
            if c:
                cleaned.append(c)
        return cleaned

    def extract_lecturer_info(self, cell_content):
        """Extract lecturer information from cell content - handles ALL formats"""
        if not cell_content or cell_content in ["FREE", "BREAK", "", "Free"]:
            return None

        # Format 1: Single-line labeled (common)
        # "Course: PHY 101, Lecturer: Dr. A / Dr. B, Room: RoomName"
        if 'Course:' in cell_content and 'Lecturer:' in cell_content:
            course_code = None
            room = None
            faculty = None

            course_match = re.search(r'Course:\s*(.*?)(?:,?\s*Lecturer:|\nLecturer:|$)', cell_content, re.IGNORECASE | re.DOTALL)
            lecturer_match = re.search(r'Lecturer:\s*(.*?)(?:,?\s*Room:|\nRoom:|$)', cell_content, re.IGNORECASE | re.DOTALL)
            room_match = re.search(r'Room:\s*([^\n,]+)', cell_content, re.IGNORECASE)

            if course_match:
                course_part = (course_match.group(1) or '').strip()
                if ' - ' in course_part:
                    course_code = course_part.split(' - ')[0].strip()
                else:
                    course_code = course_part.strip()

            if lecturer_match:
                faculty_part = (lecturer_match.group(1) or '').strip()
                faculty = re.sub(r'\s*\([^)]*\)\s*', '', faculty_part).strip()

            if room_match:
                room = (room_match.group(1) or '').strip()

            if course_code:
                return {
                    'course_code': course_code,
                    'room': room or "",
                    'faculty': faculty or ""
                }
        
        # Format 2: Newline-separated with labels
        # Course: GST 111\nLecturer: Dr. Name\nRoom: RoomName
        lines = cell_content.split('\n')
        course_code = None
        room = None
        faculty = None
        
        for line in lines:
            line = line.strip()
            if line.startswith('Course:'):
                course_part = line.replace('Course:', '').strip()
                if ' - ' in course_part:
                    course_code = course_part.split(' - ')[0].strip()
                else:
                    course_code = course_part.strip()
            elif line.startswith('Lecturer:'):
                faculty_part = line.replace('Lecturer:', '').strip()
                faculty = re.sub(r'\s*\([^)]*\)\s*', '', faculty_part).strip()
            elif line.startswith('Room:'):
                room = line.replace('Room:', '').strip()
        
        if course_code:
            return {
                'course_code': course_code,
                'room': room or "",
                'faculty': faculty or ""
            }
        
        # Format 3: Old SEPT 13 format (no labels)
        # Course Code\nRoom Name\nFaculty
        if len(lines) >= 3 and not any(label in cell_content for label in ['Course:', 'Lecturer:', 'Room:']):
            return {
                'course_code': lines[0].strip(),
                'room': lines[1].strip(),
                'faculty': lines[2].strip()
            }
        
        return None
    def create_combined_program_sheet(self, wb, sheet_name, groups_data):
        """Create a combined timetable sheet for multiple student groups of the same program"""
        ws = wb.create_sheet(title=sheet_name)
        
        current_row = 1
        
        for group_idx, group_data in enumerate(groups_data):
            full_group_name = group_data['student_group']['name']
            
            # Add student group name header
            group_header_cell = ws.cell(row=current_row, column=1, value=full_group_name)
            group_header_cell.font = Font(bold=True, size=14)
            group_header_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Create headers - Fixed order: Course Code, Course Name, Units, then TIME, DAY, CLASSROOM pattern
            headers = [
                ("Course Code", "A"),
                ("Course Name", "B"), 
                ("Units", "C"),
                ("TIME", "D"),
                ("MONDAY", "E"),
                ("CLASSROOM", "F"),
                ("TIME", "G"),
                ("TUESDAY", "H"),
                ("CLASSROOM", "I"),
                ("TIME", "J"),
                ("WEDNESDAY", "K"),
                ("CLASSROOM", "L"),
                ("TIME", "M"),
                ("THURSDAY", "N"),
                ("CLASSROOM", "O"),
                ("TIME", "P"),
                ("FRIDAY", "Q"),
                ("CLASSROOM", "R")
            ]
            
            # Set headers and apply formatting
            for col_idx, (header_text, _) in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply green color to Time and day columns
                if header_text in ["TIME", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]:
                    cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                
                # Apply borders to headers from TIME column onwards (column 4+)
                if col_idx >= 4:  # TIME column and onwards
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
            
            current_row += 1
            
            # Extract unique courses from timetable
            courses_data = self.extract_courses_from_timetable(group_data['timetable'])
            
            # Fill time slots and schedule data
            for time_idx in range(len(self.time_slots)):
                time_row = current_row + time_idx
                
                # Define column groups for each day (TIME, DAY, CLASSROOM)
                day_column_groups = [
                    (4, 5, 6),   # Monday
                    (7, 8, 9),   # Tuesday
                    (10, 11, 12), # Wednesday
                    (13, 14, 15), # Thursday
                    (16, 17, 18)  # Friday
                ]
                
                for day_idx, (time_col, day_col, classroom_col) in enumerate(day_column_groups):
                    # Time column
                    time_cell = ws.cell(row=time_row, column=time_col, value=self.time_slots[time_idx])
                    time_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                    time_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Check if this is lunch break time (1:00-1:50) and specific days
                    is_lunch_break = (time_idx == 4 and day_idx in [0, 2, 4])  # Monday(0), Wednesday(2), Friday(4)
                    
                    if is_lunch_break:
                        # Day column (BREAK)
                        day_cell = ws.cell(row=time_row, column=day_col, value="BREAK")
                        day_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Empty classroom and building for break
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value="")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                    else:
                        # Get class info for this time and day
                        class_info = self.get_class_at_time_day(group_data['timetable'], time_idx, day_idx)
                        
                        # Day column (course code)
                        day_cell = ws.cell(row=time_row, column=day_col, value=class_info['course_code'] if class_info else "")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Classroom column
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value=class_info['room'] if class_info else "")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                    # Apply borders to all timetable cells (from TIME column onwards)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    time_cell.border = border
                    day_cell.border = border
                    classroom_cell.border = border
            
            # Fill course information (A, B, C columns) - NO BORDERS for these
            course_row = current_row
            for course_code, course_info in courses_data.items():
                # Course Code (A)
                code_cell = ws.cell(row=course_row, column=1, value=course_code)
                code_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Course Name (B)
                course_name = self.get_course_name(course_code)
                name_cell = ws.cell(row=course_row, column=2, value=course_name)
                name_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Units (C)
                units = course_info.get('hours', 1)
                units_cell = ws.cell(row=course_row, column=3, value=units)
                units_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                course_row += 1
            
            # Move to next student group (4 lines below)
            current_row += len(self.time_slots) + 4
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width

    def create_student_group_timetable_sheet(self, wb, sheet_name, group_data):
        """Create a timetable sheet for a student group"""
        ws = wb.create_sheet(title=sheet_name)
        
        # Create header row with TIME and days
        headers = ["TIME"] + self.days
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Fill in the timetable data
        timetable_rows = group_data['timetable']
        
        for row_idx, row_data in enumerate(timetable_rows):
            excel_row = row_idx + 2  # Start from row 2 (after header)
            
            # Add time slot
            time_cell = ws.cell(row=excel_row, column=1, value=self.time_slots[row_idx])
            time_cell.font = Font(bold=True, color="FFFFFF")
            time_cell.fill = PatternFill(start_color="11214D", end_color="11214D", fill_type="solid")
            time_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add class data for each day
            for day_idx in range(len(self.days)):
                col = day_idx + 2  # Start from column 2 (after TIME column)
                
                if day_idx + 1 < len(row_data):  # Skip time column in source data
                    cell_content = row_data[day_idx + 1]
                    
                    if cell_content and cell_content not in ["FREE", ""]:
                        # Parse cell content
                        lines = cell_content.split('\n')
                        if len(lines) >= 3:
                            course_code = lines[0]
                            room = lines[1]
                            faculty = lines[2]
                            
                            # Format as: Course Code, Room, Faculty
                            display_text = f"{course_code}\n{room}\n{faculty}"
                        else:
                            display_text = cell_content
                    else:
                        display_text = ""
                    
                    cell = ws.cell(row=excel_row, column=col, value=display_text)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Apply border
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column].width = adjusted_width
        
        # Set row heights for better visibility
        for row in range(1, len(timetable_rows) + 2):
            ws.row_dimensions[row].height = 60

    def create_combined_lecturer_sheet(self, wb, sheet_name, lecturers_data):
        """Create a combined timetable sheet for all lecturers"""
        ws = wb.create_sheet(title=sheet_name)
        
        current_row = 1
        
        for lecturer_name, lecturer_schedule in lecturers_data.items():
            # Add lecturer name header
            lecturer_header_cell = ws.cell(row=current_row, column=1, value=lecturer_name)
            lecturer_header_cell.font = Font(bold=True, size=14)
            lecturer_header_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Create headers - Fixed order: Course Code, Course Name, Units, then TIME, DAY, CLASSROOM pattern
            headers = [
                ("Course Code", "A"),
                ("Course Name", "B"), 
                ("Units", "C"),
                ("TIME", "D"),
                ("MONDAY", "E"),
                ("CLASSROOM", "F"),
                ("TIME", "G"),
                ("TUESDAY", "H"),
                ("CLASSROOM", "I"),
                ("TIME", "J"),
                ("WEDNESDAY", "K"),
                ("CLASSROOM", "L"),
                ("TIME", "M"),
                ("THURSDAY", "N"),
                ("CLASSROOM", "O"),
                ("TIME", "P"),
                ("FRIDAY", "Q"),
                ("CLASSROOM", "R")
            ]
            
            # Set headers and apply formatting
            for col_idx, (header_text, _) in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header_text)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply green color to Time and day columns
                if header_text in ["TIME", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]:
                    cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                
                # Apply borders to headers from TIME column onwards (column 4+)
                if col_idx >= 4:  # TIME column and onwards
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
            
            current_row += 1
            
            # Extract unique courses from lecturer schedule
            courses_data = self.extract_courses_from_lecturer_schedule(lecturer_schedule)
            
            # Create schedule grid
            schedule_grid = {}
            for time_idx in range(len(self.time_slots)):
                schedule_grid[time_idx] = {}
                for day_idx in range(len(self.days)):
                    schedule_grid[time_idx][day_idx] = None
            
            # Fill lecturer's schedule
            for class_info in lecturer_schedule:
                time_idx = class_info['time_slot']
                day_idx = class_info['day']
                schedule_grid[time_idx][day_idx] = class_info
            
            # Fill time slots and schedule data
            for time_idx in range(len(self.time_slots)):
                time_row = current_row + time_idx
                
                # Define column groups for each day (TIME, DAY, CLASSROOM)
                day_column_groups = [
                    (4, 5, 6),   # Monday
                    (7, 8, 9),   # Tuesday
                    (10, 11, 12), # Wednesday
                    (13, 14, 15), # Thursday
                    (16, 17, 18)  # Friday
                ]
                
                for day_idx, (time_col, day_col, classroom_col) in enumerate(day_column_groups):
                    # Time column
                    time_cell = ws.cell(row=time_row, column=time_col, value=self.time_slots[time_idx])
                    time_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                    time_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Check if this is lunch break time (1:00-1:50) and specific days
                    is_lunch_break = (time_idx == 4 and day_idx in [0, 2, 4])  # Monday(0), Wednesday(2), Friday(4)
                    
                    if is_lunch_break:
                        # Day column (BREAK)
                        day_cell = ws.cell(row=time_row, column=day_col, value="BREAK")
                        day_cell.fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Empty classroom and building for break
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value="")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")

                    else:
                        # Get class info for this time and day
                        class_info = schedule_grid[time_idx][day_idx]
                        
                        # Day column (course code)
                        day_cell = ws.cell(row=time_row, column=day_col, value=class_info['course_code'] if class_info else "")
                        day_cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Classroom column
                        classroom_cell = ws.cell(row=time_row, column=classroom_col, value=class_info['room'] if class_info else "")
                        classroom_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Apply borders to all timetable cells (from TIME column onwards)
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    time_cell.border = border
                    day_cell.border = border
                    classroom_cell.border = border
            
            # Fill course information (A, B, C columns) - NO BORDERS for these
            course_row = current_row
            for course_code, course_info in courses_data.items():
                # Course Code (A)
                code_cell = ws.cell(row=course_row, column=1, value=course_code)
                code_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Course Name (B)
                course_name = self.get_course_name(course_code)
                name_cell = ws.cell(row=course_row, column=2, value=course_name)
                name_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Units (C)
                units = course_info.get('hours', 1)
                units_cell = ws.cell(row=course_row, column=3, value=units)
                units_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                course_row += 1
            
            # Move to next lecturer (4 lines below)
            current_row += len(self.time_slots) + 4
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = None
            
            # Find the first non-merged cell to get column letter
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break
            
            if column:
                for cell in col:
                    try:
                        if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width

    def export_sst_timetables(self):
        """Export SST (engineering) timetables"""
        data = self.get_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Group SST student groups by main program
            sst_programs = defaultdict(list)
            
            for group_data in data:
                sg_obj = group_data.get('student_group') or {}
                group_name = sg_obj.get('name') if isinstance(sg_obj, dict) else str(sg_obj)
                if self.is_sst_group(sg_obj):
                    main_program = self.extract_main_program_name(group_name)
                    sst_programs[main_program].append(group_data)
            
            if not sst_programs:
                return False, "No SST student groups found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each unique program (one sheet per main program)
            for program_name, groups in sst_programs.items():
                # Sort groups by year for better organization
                groups.sort(key=lambda x: x['student_group']['name'])
                
                # Use main program name as sheet name, but limit length for Excel
                safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
                self.create_combined_program_sheet(wb, safe_sheet_name, groups)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SST_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"SST Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting SST timetables: {str(e)}"

    def export_tyd_timetables(self):
        """Export TYD (non-engineering) timetables"""
        data = self.get_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Group TYD student groups by main program
            tyd_programs = defaultdict(list)
            
            for group_data in data:
                sg_obj = group_data.get('student_group') or {}
                group_name = sg_obj.get('name') if isinstance(sg_obj, dict) else str(sg_obj)
                if not self.is_sst_group(sg_obj):  # Not SST = TYD
                    main_program = self.extract_main_program_name(group_name)
                    tyd_programs[main_program].append(group_data)
            
            if not tyd_programs:
                return False, "No TYD student groups found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets for each unique program (one sheet per main program)
            for program_name, groups in tyd_programs.items():
                # Sort groups by year for better organization
                groups.sort(key=lambda x: x['student_group']['name'])
                
                # Use main program name as sheet name, but limit length for Excel
                safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
                self.create_combined_program_sheet(wb, safe_sheet_name, groups)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"TYD_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"TYD Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting TYD timetables: {str(e)}"

    def export_lecturer_timetables(self):
        """Export all lecturer timetables"""
        data = self.get_timetable_data()
        if not data:
            return False, "No timetable data available"
        
        try:
            # Collect all lecturer schedules
            lecturer_schedules = defaultdict(list)
            
            for group_data in data:
                student_group_name = group_data['student_group']['name']
                timetable_rows = group_data['timetable']
                
                for time_slot_idx, row_data in enumerate(timetable_rows):
                    for day_idx in range(len(self.days)):
                        if day_idx + 1 < len(row_data):  # Skip time column
                            cell_content = row_data[day_idx + 1]
                            lecturer_info = self.extract_lecturer_info(cell_content)
                            
                            if lecturer_info and lecturer_info['course_code']:
                                course_code = lecturer_info['course_code']

                                room = lecturer_info.get('room', '')
                                faculty_raw = lecturer_info.get('faculty', '')
                                lecturer_names = self.split_lecturer_names(faculty_raw)

                                # Fallback: if Lecturer wasn't present in the cell, use course_data mappings.
                                if not lecturer_names:
                                    course_info = self.course_data.get(course_code, {})
                                    faculty_ids = course_info.get('facultyId', [])
                                    if isinstance(faculty_ids, str):
                                        faculty_ids = [faculty_ids]
                                    for faculty_id in faculty_ids:
                                        faculty_info = self.faculty_data.get(faculty_id, {})
                                        name = faculty_info.get('name', faculty_id)
                                        if name and name.lower() not in ['unknown', 'tbd', 'staff']:
                                            lecturer_names.append(str(name).strip())

                                for lecturer_name in lecturer_names:
                                    if not lecturer_name or lecturer_name.lower() in ['unknown', 'tbd', 'staff', '']:
                                        continue
                                    lecturer_schedules[lecturer_name].append({
                                        'time_slot': time_slot_idx,
                                        'day': day_idx,
                                        'course_code': course_code,
                                        'room': room,
                                        'student_group': student_group_name
                                    })
            
            if not lecturer_schedules:
                return False, "No lecturer data found"
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create one combined sheet for all lecturers
            self.create_combined_lecturer_sheet(wb, "Lecturer Timetables", lecturer_schedules)
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Lecturer_Timetables_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            return True, f"Lecturer Timetables exported successfully to {filename}"
            
        except Exception as e:
            return False, f"Error exporting lecturer timetables: {str(e)}"

    def extract_courses_from_timetable(self, timetable_rows):
        """Extract unique courses from timetable data"""
        courses = {}
        
        for row_idx, row_data in enumerate(timetable_rows):
            for day_idx in range(len(self.days)):
                if day_idx + 1 < len(row_data):  # Skip time column
                    cell_content = row_data[day_idx + 1]
                    
                    if cell_content and cell_content not in ["FREE", "", "BREAK", "Free"]:
                        # Use the updated parser that handles all formats
                        class_info = self.extract_lecturer_info(cell_content)
                        if class_info and class_info['course_code']:
                            course_code = class_info['course_code']
                            if course_code not in courses:
                                courses[course_code] = {'hours': 0}
                            courses[course_code]['hours'] += 1
        
        return courses

    def extract_courses_from_lecturer_schedule(self, lecturer_schedule):
        """Extract unique courses from lecturer schedule data"""
        courses = {}
        
        for class_info in lecturer_schedule:
            course_code = class_info['course_code']
            if course_code:
                if course_code not in courses:
                    courses[course_code] = {'hours': 0}
                courses[course_code]['hours'] += 1
        
        return courses

    def get_class_at_time_day(self, timetable_rows, time_idx, day_idx):
        """Get class information at specific time and day"""
        if time_idx < len(timetable_rows):
            row_data = timetable_rows[time_idx]
            if day_idx + 1 < len(row_data):  # Skip time column
                cell_content = row_data[day_idx + 1]
                
                if cell_content and cell_content not in ["FREE", "", "BREAK", "Free"]:
                    # Use the updated parser that handles all formats
                    return self.extract_lecturer_info(cell_content)
        return None

    def get_building_from_room(self, room_name):
        """Determine building from room data"""
        if not room_name:
            return ""
        
        # Look up room in the loaded room data
        room_info = self.room_data.get(room_name)
        if room_info and 'building' in room_info:
            return room_info['building']
        
        # Fallback to heuristic if not found in data
        room_lower = room_name.lower()
        if any(keyword in room_lower for keyword in ['eng', 'lab', 'workshop', 'tech']):
            return "SST"
        else:
            return "TYD"

    def get_course_name(self, course_code):
        """Get full course name from course code using loaded course data"""
        if not course_code:
            return ""
        
        # Look up course in the loaded course data
        course_info = self.course_data.get(course_code)
        if course_info and 'name' in course_info:
            return course_info['name']
        
        # Fallback if not found in data
        return f"{course_code} (Course Name)"

# Main functions for easy import
def export_sst_timetables():
    """Export SST timetables"""
    exporter = TimetableExporter()
    return exporter.export_sst_timetables()

def export_tyd_timetables():
    """Export TYD timetables"""
    exporter = TimetableExporter()
    return exporter.export_tyd_timetables()

def export_lecturer_timetables():
    """Export lecturer timetables"""
    exporter = TimetableExporter()
    return exporter.export_lecturer_timetables()

# New functions that accept data directly from Dash UI and return bytes for browser downloads
def export_sst_timetables_bytes_from_data(timetable_data):
    """Export SST timetables from Dash UI data and return bytes for download"""
    from io import BytesIO
    
    if not timetable_data:
        return None, "No timetable data available"
    
    exporter = TimetableExporter()
    
    # Extract timetables list
    data = timetable_data.get('timetables', []) if isinstance(timetable_data, dict) else timetable_data
    
    if not data:
        return None, "No timetable data available"
    
    try:
        # Group SST student groups by main program
        sst_programs = defaultdict(list)
        
        for group_data in data:
            sg_obj = group_data.get('student_group') or {}
            group_name = sg_obj.get('name') if isinstance(sg_obj, dict) else str(sg_obj)
            if exporter.is_sst_group(sg_obj):
                main_program = exporter.extract_main_program_name(group_name)
                sst_programs[main_program].append(group_data)
        
        if not sst_programs:
            return None, "No SST student groups found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for program_name, groups in sst_programs.items():
            groups.sort(key=lambda x: x['student_group']['name'])
            safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
            exporter.create_combined_program_sheet(wb, safe_sheet_name, groups)
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SST_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        import traceback
        error_msg = f"Error: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        return None, error_msg

def export_tyd_timetables_bytes_from_data(timetable_data):
    """Export TYD timetables from Dash UI data and return bytes for download"""
    from io import BytesIO
    
    if not timetable_data:
        return None, "No timetable data available"
    
    exporter = TimetableExporter()
    
    # Extract timetables list
    data = timetable_data.get('timetables', []) if isinstance(timetable_data, dict) else timetable_data
    
    if not data:
        return None, "No timetable data available"
    
    try:
        # Group TYD student groups by main program
        tyd_programs = defaultdict(list)
        
        for group_data in data:
            sg_obj = group_data.get('student_group') or {}
            group_name = sg_obj.get('name') if isinstance(sg_obj, dict) else str(sg_obj)
            if not exporter.is_sst_group(sg_obj):
                main_program = exporter.extract_main_program_name(group_name)
                tyd_programs[main_program].append(group_data)
        
        if not tyd_programs:
            return None, "No TYD student groups found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        for program_name, groups in tyd_programs.items():
            groups.sort(key=lambda x: x['student_group']['name'])
            safe_sheet_name = re.sub(r'[^\w\s-]', '', program_name)[:31]
            exporter.create_combined_program_sheet(wb, safe_sheet_name, groups)
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TYD_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        import traceback
        error_msg = f"Error: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        return None, error_msg

def export_lecturer_timetables_bytes_from_data(timetable_data):
    """Export lecturer timetables from Dash UI data and return bytes for download"""
    from io import BytesIO
    
    if not timetable_data:
        return None, "No timetable data available"
    
    exporter = TimetableExporter()
    
    # Extract timetables list
    data = timetable_data.get('timetables', []) if isinstance(timetable_data, dict) else timetable_data
    
    if not data:
        return None, "No timetable data available"
    
    try:
        # Collect all lecturer schedules
        lecturer_schedules = defaultdict(list)
        
        for group_data in data:
            student_group_name = group_data['student_group']['name']
            timetable_rows = group_data['timetable']
            
            for time_slot_idx, row_data in enumerate(timetable_rows):
                for day_idx in range(len(exporter.days)):
                    if day_idx + 1 < len(row_data):  # Skip time column
                        cell_content = row_data[day_idx + 1]
                        lecturer_info = exporter.extract_lecturer_info(cell_content)
                        
                        if lecturer_info and lecturer_info['course_code']:
                            course_code = lecturer_info['course_code']

                            room = lecturer_info.get('room', '')
                            faculty_raw = lecturer_info.get('faculty', '')
                            lecturer_names = exporter.split_lecturer_names(faculty_raw)

                            if not lecturer_names:
                                course_info = exporter.course_data.get(course_code, {})
                                faculty_ids = course_info.get('facultyId', [])
                                if isinstance(faculty_ids, str):
                                    faculty_ids = [faculty_ids]
                                for faculty_id in faculty_ids:
                                    faculty_info = exporter.faculty_data.get(faculty_id, {})
                                    name = faculty_info.get('name', faculty_id)
                                    if name and name.lower() not in ['unknown', 'tbd', 'staff']:
                                        lecturer_names.append(str(name).strip())

                            for lecturer_name in lecturer_names:
                                if not lecturer_name or lecturer_name.lower() in ['unknown', 'tbd', 'staff', '']:
                                    continue
                                lecturer_schedules[lecturer_name].append({
                                    'time_slot': time_slot_idx,
                                    'day': day_idx,
                                    'course_code': course_code,
                                    'room': room,
                                    'student_group': student_group_name
                                })
        
        if not lecturer_schedules:
            return None, "No lecturer data found"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        exporter.create_combined_lecturer_sheet(wb, "Lecturer Timetables", lecturer_schedules)
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Lecturer_Timetables_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
        
    except Exception as e:
        import traceback
        error_msg = f"Error: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        return None, error_msg


def export_classrooms_scheduled_bytes_from_data(timetables: list, rooms: list, upload_id: str = ""):
    """Export a 'Classrooms Scheduled' workbook.

    Output:
    - Two sheets: 'TYD Classrooms' and 'SST Classrooms'
    - Each sheet contains ALL rooms from that building (even unused)
    - Rooms are sorted by usage (most-used first)
    - For each room: days are laid out horizontally (Mon-Fri) with columns:
        Student Group | Course | Times used
      where Times used merges consecutive slots into ranges (comma-separated).
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    day_names = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

    def _norm_building(v) -> str:
        s = str(v or "").strip().upper()
        if s.startswith("SST"):
            return "SST"
        if s.startswith("TYD"):
            return "TYD"
        return s

    def _room_name(room) -> str:
        if isinstance(room, dict):
            return str(room.get("name") or room.get("Id") or room.get("id") or "").strip()
        return str(getattr(room, "name", "") or getattr(room, "Id", "") or getattr(room, "id", "") or "").strip()

    def _room_building(room, room_name: str) -> str:
        building = ""
        if isinstance(room, dict):
            building = _norm_building(room.get("building"))
        else:
            building = _norm_building(getattr(room, "building", None))

        if building in {"SST", "TYD"}:
            return building

        n = (room_name or "").upper()
        if "SST" in n:
            return "SST"
        if "TYD" in n:
            return "TYD"

        # Default unknown rooms to TYD to ensure they're not dropped.
        return "TYD"

    def _fmt_hhmm(total_minutes: int) -> str:
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h}:{m:02d}"

    def _slot_range_for_row_idx(row_idx: int) -> tuple[int, int]:
        start = (8 * 60 + 30) + row_idx * 60
        end = start + 60
        return start, end

    def _merge_slot_indices(slot_indices: list[int]) -> str:
        if not slot_indices:
            return ""
        slot_indices = sorted(set(int(x) for x in slot_indices))
        ranges: list[str] = []
        start_idx = slot_indices[0]
        end_idx = slot_indices[0]
        for idx in slot_indices[1:]:
            if idx == end_idx + 1:
                end_idx = idx
                continue
            start_min, _ = _slot_range_for_row_idx(start_idx)
            _, end_min = _slot_range_for_row_idx(end_idx)
            ranges.append(f"{_fmt_hhmm(start_min)}-{_fmt_hhmm(end_min)}")
            start_idx = end_idx = idx
        start_min, _ = _slot_range_for_row_idx(start_idx)
        _, end_min = _slot_range_for_row_idx(end_idx)
        ranges.append(f"{_fmt_hhmm(start_min)}-{_fmt_hhmm(end_min)}")
        return ", ".join(ranges)

    def _parse_cell(cell: str) -> tuple[str, str]:
        """Return (course_code, room_name) from a cell string."""
        text = str(cell or "").strip()
        if not text or text.upper() in {"FREE", "BREAK"}:
            return "", ""
        course = ""
        room = ""
        m = re.search(r"Course:\s*([^,\n]+)", text, flags=re.IGNORECASE)
        if m:
            course = m.group(1).strip()
        m = re.search(r"Room:\s*([^\n,]+)", text, flags=re.IGNORECASE)
        if m:
            room = m.group(1).strip()
        return course, room

    # ---- Build room lists (include all rooms, even unused) ----
    rooms_by_building: dict[str, list[str]] = {"TYD": [], "SST": []}
    for room in (rooms or []):
        rn = _room_name(room)
        if not rn:
            continue
        b = _room_building(room, rn)
        rooms_by_building.setdefault(b, [])
        rooms_by_building[b].append(rn)

    for b in ["TYD", "SST"]:
        seen: set[str] = set()
        uniq: list[str] = []
        for rn in rooms_by_building.get(b, []):
            key = rn.strip().lower()
            if key and key not in seen:
                seen.add(key)
                uniq.append(rn)
        rooms_by_building[b] = uniq

    # ---- Index usage by room/day/(sg,course) -> slot indices ----
    usage: dict[str, dict[int, dict[tuple[str, str], list[int]]]] = {}
    usage_count: dict[str, int] = {}

    for entry in (timetables or []):
        sg = (entry.get("student_group") or {}) if isinstance(entry, dict) else {}
        if isinstance(sg, dict):
            sg_name = str(sg.get("name") or sg.get("id") or "Unknown").strip()
        else:
            sg_name = str(sg or "Unknown").strip()
        grid = (entry.get("timetable") or entry.get("rows") or []) if isinstance(entry, dict) else []

        for row_idx, row in enumerate(grid):
            if not isinstance(row, list) or len(row) < 2:
                continue
            for day_idx, cell in enumerate((row[1:] or [])[:5]):
                course, room = _parse_cell(cell)
                if not room or not course:
                    continue
                usage.setdefault(room, {}).setdefault(day_idx, {}).setdefault((sg_name, course), []).append(row_idx)
                usage_count[room] = usage_count.get(room, 0) + 1

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    # Style (match the green header scheme in your example)
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Use the same green scheme as the timetable headers (no dark-blue bars)
    room_header_fill = PatternFill(start_color="FFFFFF", end_color="BDEDBD", fill_type="solid")
    room_header_font = Font(color="000000", bold=True, size=14)

    day_header_fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
    day_header_font = Font(bold=True)

    sub_header_fill = PatternFill(start_color="BDEDBD", end_color="BDEDBD", fill_type="solid")
    sub_header_font = Font(bold=True)

    def _apply_block_border(ws, start_row: int, end_row: int, start_col: int, end_col: int):
        for rr in range(start_row, end_row + 1):
            for cc in range(start_col, end_col + 1):
                ws.cell(row=rr, column=cc).border = border

    def write_sheet(building: str, title: str):
        ws = wb.create_sheet(title)

        cols_per_day = 3
        total_cols = len(day_names) * cols_per_day

        for d in range(len(day_names)):
            base = 1 + d * cols_per_day
            ws.column_dimensions[get_column_letter(base)].width = 22  # Student Group
            ws.column_dimensions[get_column_letter(base + 1)].width = 14  # Course
            ws.column_dimensions[get_column_letter(base + 2)].width = 18  # Times used

        room_list = rooms_by_building.get(building, [])
        room_list = sorted(room_list, key=lambda rn: (-usage_count.get(rn, 0), rn.lower()))

        r = 1
        for rn in room_list:
            header_row = r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
            cell = ws.cell(row=r, column=1, value=f"CLASSROOM: {rn}    (Used slots: {usage_count.get(rn, 0)})")
            cell.fill = room_header_fill
            cell.font = room_header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            r += 1

            day_row = r
            for day_idx, day in enumerate(day_names):
                start_col = 1 + day_idx * cols_per_day
                end_col = start_col + cols_per_day - 1
                ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
                dcell = ws.cell(row=r, column=start_col, value=day)
                dcell.fill = day_header_fill
                dcell.font = day_header_font
                dcell.alignment = Alignment(horizontal="center", vertical="center")
            r += 1

            sub_row = r
            for day_idx in range(len(day_names)):
                base = 1 + day_idx * cols_per_day
                for off, text in enumerate(["Student Group", "Course", "Times used"]):
                    hcell = ws.cell(row=r, column=base + off, value=text)
                    hcell.fill = sub_header_fill
                    hcell.font = sub_header_font
                    hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            r += 1

            day_items: list[list[tuple[str, str, str]]] = []
            max_rows = 0
            for day_idx in range(len(day_names)):
                day_map = usage.get(rn, {}).get(day_idx, {})
                if not day_map:
                    items = [("No classes scheduled", "", "")]
                else:
                    items = [(sgn, c, _merge_slot_indices(idxs)) for (sgn, c), idxs in day_map.items()]
                    items.sort(key=lambda x: (x[0].lower(), x[1].lower()))
                day_items.append(items)
                max_rows = max(max_rows, len(items))

            data_start_row = r
            for i in range(max_rows):
                for day_idx in range(len(day_names)):
                    base = 1 + day_idx * cols_per_day
                    items = day_items[day_idx]
                    sgn, c, tstr = items[i] if i < len(items) else ("", "", "")
                    sg_cell = ws.cell(row=r, column=base, value=sgn)
                    sg_cell.font = Font(bold=True)
                    sg_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                    c_cell = ws.cell(row=r, column=base + 1, value=c)
                    c_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

                    t_cell = ws.cell(row=r, column=base + 2, value=tstr)
                    t_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                r += 1

            _apply_block_border(ws, header_row, r - 1, 1, total_cols)
            r += 2

        if ws.max_row >= 3:
            ws.freeze_panes = "A4"

        return ws

    write_sheet("TYD", "TYD Classrooms")
    write_sheet("SST", "SST Classrooms")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Classrooms_Scheduled_{upload_id or 'export'}.xlsx"
    return buffer.getvalue(), filename

if __name__ == "__main__":
    # Test the export functions
    print("Testing SST export...")
    success, message = export_sst_timetables()
    print(f"SST: {message}")
    
    print("\nTesting TYD export...")
    success, message = export_tyd_timetables()
    print(f"TYD: {message}")
    
    print("\nTesting Lecturer export...")
    success, message = export_lecturer_timetables()
    print(f"Lecturer: {message}")
