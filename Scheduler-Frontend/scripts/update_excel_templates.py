from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


TEMPLATES = [
    Path("public/Timetable_Input_Template.xlsx"),
    Path("Input System Details/Timetable_Input_Template.xlsx"),
    Path("PAU-Timetable-Scheduler/data/Timetable_Input_Template.xlsx"),
    Path("PAU-Timetable-Scheduler/data/NEWLY updated Timetable_Input_Template.xlsx"),
]


def _norm(value: object) -> str:
    return str(value or "").strip().lower()


def _find_lecturers_sheet(wb):
    for name in wb.sheetnames:
        if _norm(name) in {
            "lecturers",
            "lecturer",
            "faculty",
            "faculties",
            "teachers",
            "instructors",
        }:
            return wb[name]
    if "Lecturers" in wb.sheetnames:
        return wb["Lecturers"]
    return None


def ensure_columns(path: Path) -> str:
    if not path.exists():
        return f"SKIP missing: {path}"

    wb = load_workbook(path)
    sheet = _find_lecturers_sheet(wb)
    if sheet is None:
        return f"WARN no Lecturers sheet: {path}"

    headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
    headers_norm = [_norm(h) for h in headers]

    if "available hours" in headers_norm and "available consecutive hours" in headers_norm:
        wb.save(path)
        return f"OK already has columns: {path}"

    try:
        avail_times_col = headers_norm.index("available times") + 1
    except ValueError:
        avail_times_col = sheet.max_column

    insert_at = avail_times_col + 1
    sheet.insert_cols(insert_at, amount=2)
    sheet.cell(row=1, column=insert_at).value = "Available Hours"
    sheet.cell(row=1, column=insert_at + 1).value = "Available Consecutive Hours"

    wb.save(path)
    return f"UPDATED: {path}"


def main() -> None:
    for path in TEMPLATES:
        print(ensure_columns(path))


if __name__ == "__main__":
    main()
